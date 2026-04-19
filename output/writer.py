import os
from datetime import date
from typing import Optional, TypedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class AnalysisRow(TypedDict):
    # --- Product info ---
    gtin: str
    product_name: Optional[str]
    cost_price: float
    your_qogita_price: float
    current_profit_eur: Optional[float]
    current_profit_pct: Optional[float]

    # --- Seller data ---
    seller1_name: Optional[str]
    seller1_max_price: Optional[float]
    seller1_stock: Optional[int]
    seller1_delivery: Optional[str]
    seller2_name: Optional[str]
    seller2_max_price: Optional[float]
    seller2_stock: Optional[int]

    # --- Scenario A: be cheapest seller (always shown) ---
    a_suggested_price: Optional[float]
    a_profit_eur: Optional[float]
    a_profit_pct: Optional[float]
    a_will_be_cheapest: Optional[bool]
    a_notes: str

    # --- Scenario B: seller 1 has slow delivery ---
    b_suggested_price: Optional[float]
    b_profit_eur: Optional[float]
    b_profit_pct: Optional[float]
    b_delivery_info: Optional[str]
    b_notes: str

    # --- Scenario C: seller 1 low stock, wait them out ---
    c_suggested_price: Optional[float]
    c_profit_eur: Optional[float]
    c_profit_pct: Optional[float]
    c_price_gap_pct: Optional[float]
    c_notes: str


# (header text, background fill hex or None)
HEADERS: list[tuple[str, Optional[str]]] = [
    # Product info — light grey
    ("GTIN",                          "D9D9D9"),
    ("Product Name",                  "D9D9D9"),
    ("Cost Price",                    "D9D9D9"),
    ("Your Current Price",            "D9D9D9"),
    ("Current Profit €",              "D9D9D9"),
    ("Current Profit %",              "D9D9D9"),

    # Seller data — white
    ("Seller 1",                      None),
    ("Seller 1 Price (excl. margin)",  None),
    ("Seller 1 Stock",                None),
    ("Seller 1 Delivery",             None),
    ("Seller 2",                      None),
    ("Seller 2 Price (excl. margin)",  None),
    ("Seller 2 Stock",                None),

    # Scenario A — light blue
    ("A: Suggested Price (Cheapest)", "BDD7EE"),
    ("A: Profit €",                   "BDD7EE"),
    ("A: Profit %",                   "BDD7EE"),
    ("A: Will You Be Cheapest?",      "BDD7EE"),
    ("A: Notes",                      "BDD7EE"),

    # Scenario B — light orange
    ("B: Suggested Price (2nd Best — Long Delivery)", "FCE4D6"),
    ("B: Profit €",                   "FCE4D6"),
    ("B: Profit %",                   "FCE4D6"),
    ("B: Seller 1 Delivery",          "FCE4D6"),
    ("B: Notes",                      "FCE4D6"),

    # Scenario C — light yellow
    ("C: Suggested Price (2nd Best — Low Stock)", "FFEB9C"),
    ("C: Profit €",                   "FFEB9C"),
    ("C: Profit %",                   "FFEB9C"),
    ("C: Seller 1 Stock Gap %",       "FFEB9C"),
    ("C: Notes",                      "FFEB9C"),
]

# Column index (1-based) for the A: Notes column — used for row colour coding
A_NOTES_COL = next(i + 1 for i, (h, _) in enumerate(HEADERS) if h == "A: Notes")

NOTE_COLORS = {
    "Lower needed":            "FFFFC0",  # yellow
    "Price increase possible": "CCE5FF",  # blue
    "Optimal price":           "CCFFCC",  # green
    "Can't compete":           "FFCCCC",  # red
    "Not found":               "E0E0E0",  # grey
    "Extraction failed":       "FFD9B3",  # orange
    "No current price set":    "E0E0E0",  # grey
}


def _fmt_bool(val: Optional[bool]) -> Optional[str]:
    if val is None:
        return None
    return "Yes" if val else "No"


def _fmt_pct(val: Optional[float]) -> Optional[str]:
    if val is None:
        return None
    return f"{val}%"


def write_excel(rows: list[AnalysisRow], output_dir: str = "output") -> str:
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{date.today().strftime('%Y-%m-%d')}_qogita_products_analysis.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # Header row
    for col, (header, fill_hex) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if fill_hex:
            cell.fill = PatternFill("solid", fgColor=fill_hex)

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        a_notes = row.get("a_notes") or ""

        values = [
            # Product info
            row["gtin"],
            row.get("product_name"),
            row["cost_price"],
            row["your_qogita_price"],
            row.get("current_profit_eur"),
            _fmt_pct(row.get("current_profit_pct")),
            # Seller data
            row.get("seller1_name"),
            row.get("seller1_max_price"),
            row.get("seller1_stock"),
            row.get("seller1_delivery") or "In stock",
            row.get("seller2_name"),
            row.get("seller2_max_price"),
            row.get("seller2_stock"),
            # Scenario A
            row.get("a_suggested_price"),
            row.get("a_profit_eur"),
            _fmt_pct(row.get("a_profit_pct")),
            _fmt_bool(row.get("a_will_be_cheapest")),
            a_notes,
            # Scenario B
            row.get("b_suggested_price"),
            row.get("b_profit_eur"),
            _fmt_pct(row.get("b_profit_pct")),
            row.get("b_delivery_info"),
            row.get("b_notes") or "",
            # Scenario C
            row.get("c_suggested_price"),
            row.get("c_profit_eur"),
            _fmt_pct(row.get("c_profit_pct")),
            _fmt_pct(row.get("c_price_gap_pct")),
            row.get("c_notes") or "",
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=value)

        # Colour the Scenario A notes cell
        color = NOTE_COLORS.get(a_notes)
        if color:
            ws.cell(row=row_idx, column=A_NOTES_COL).fill = PatternFill("solid", fgColor=color)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    wb.save(filepath)
    return filepath
