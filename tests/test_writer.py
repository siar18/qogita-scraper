import os
from datetime import date
from output.writer import write_excel, AnalysisRow


def test_write_excel_creates_file(tmp_path):
    rows = [
        AnalysisRow(
            gtin="3614272225718",
            product_name="Giorgio Armani Stronger With You 100ml",
            your_qogita_price=65.00,
            cost_price=45.00,
            cheapest_seller="2WYZL",
            cheapest_seller_stock=20,
            cheapest_seller_max_price=60.28,
            suggested_price=53.82,
            difference=-11.18,
            notes="Lower needed"
        )
    ]
    output_path = write_excel(rows, output_dir=str(tmp_path))
    assert os.path.exists(output_path)
    assert date.today().strftime("%Y-%m-%d") in output_path


def test_write_excel_correct_columns(tmp_path):
    import openpyxl
    rows = [
        AnalysisRow(
            gtin="123",
            product_name="Test Product",
            your_qogita_price=50.0,
            cost_price=30.0,
            cheapest_seller=None,
            cheapest_seller_stock=None,
            cheapest_seller_max_price=None,
            suggested_price=None,
            difference=None,
            notes="Not found"
        )
    ]
    output_path = write_excel(rows, output_dir=str(tmp_path))
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "GTIN", "Product Name", "Your Qogita Price", "Cost Price",
        "Cheapest Seller", "Cheapest Seller Stock", "Cheapest Seller Max Price",
        "Suggested Price", "Difference", "Notes"
    ]
    assert ws.cell(2, 10).value == "Not found"
